# -*- coding: utf-8 -*-
"""
Created on Tue Dec 21 10:42:13 2021

@author: amirm
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font
import matplotlib.pyplot as plt
import numpy as np
import matplotlib.pyplot as pyplot

# Converting the extension from csv to xlsx
file_address_input = input('File address (Example: C:\Test.csv): ')
read_file = pd.read_csv(str(file_address_input))
xlsx_file = input('New file name (Example: C:\Test.xlsx): ')
read_file.to_excel(str(xlsx_file), index=None, header=True)

wb = openpyxl.load_workbook(str(xlsx_file))
ws = wb.active

# Deletion of extra columns
ws.delete_cols(1, 6)
wb.save(str(xlsx_file))
ws.delete_cols(2, 16)
wb.save(str(xlsx_file))

# Insertion of a column representing the number of distances and consequent formatting
ws.insert_cols(1, 1)
wb.save(str(xlsx_file))
ws.cell(row=1, column=1).value = 'Distance number'
ws.cell(row=1, column=2).value = 'X direction [mm]'
ws.cell(row=1, column=3).value = 'Y direction [mm]'
ws.cell(row=1, column=1).font = Font(bold=True)
ws.cell(row=1, column=3).font = Font(bold=True)
wb.save(str(xlsx_file))

# Finding the number of non-blank rows and registering the row numbers
number_of_distances = ws.max_row - 1
counter = 1
for i in range(number_of_distances):
    ws.cell(i + 2, 1).value = counter
    counter += 1
wb.save(str(xlsx_file))

# Separating of measurements in X and Y directions
section_distances = int(number_of_distances / 2)
for i in range(section_distances):
    ws.cell(i + 2, 3).value = ws.cell(section_distances + i + 2, 2).value
ws.delete_rows(section_distances + 2, section_distances)
wb.save(str(xlsx_file))

# Importing the x_direction image into the "sample_jr_report.xlsx"
x_direction_img_name = input(r'Address x_direction image (Example: C:\x_direction.bmp): ')
x_direction_img = openpyxl.drawing.image.Image(str(x_direction_img_name))
x_direction_img.anchor = 'E2'
x_direction_img.width = 600
x_direction_img.height = 500
ws.add_image(x_direction_img)
ws.cell(1, 5).value = 'X_DIRECTION'
wb.save(str(xlsx_file))

# Importing the y_direction image into the "sample_jr_report.xlsx"
y_direction_img_name = input(r'Address y_direction image (Example: C:\y_direction.bmp): ')
y_direction_img = openpyxl.drawing.image.Image(str(y_direction_img_name))
y_direction_img.anchor = 'E29'
y_direction_img.width = 600
y_direction_img.height = 500
ws.add_image(y_direction_img)
ws.cell(28, 5).value = 'Y_DIRECTION'
wb.save(str(xlsx_file))

# Showing the line graph
df = pd.read_excel(str(xlsx_file))
distance_number = list(df['Distance number'])
x_direction = list(df['X direction [mm]'])
y_direction = list(df['Y direction [mm]'])

plt.figure(figsize = (30, 10))
plt.plot(distance_number,x_direction, color = 'blue', linewidth = 2.0, label = 'X_direction', marker = 'o')
plt.plot(distance_number,y_direction, color = 'red', linewidth = 2.0, label = 'Y_direction', marker = 'o')
plt.xlabel('Number of Distances', fontsize = 'xx-large', weight = 'bold')
plt.ylabel('Distance (mm)', fontsize = 'xx-large', weight = 'bold')
plt.legend(fontsize = 'xx-large')

plt.axhline(y=min(x_direction), color = 'blue', linestyle = '--', linewidth = 1.0)
plt.annotate(('Minimum = ' + str(min(x_direction))), (distance_number[5], min(x_direction) + .01))

plt.axhline(y=max(x_direction), color = 'blue', linestyle = '--', linewidth = 1.0)
plt.annotate(('Maximum = ' + str(max(x_direction))), (distance_number[5], max(x_direction) + .01))

plt.axhline(y=min(y_direction), color = 'red', linestyle = '--', linewidth = 1.0)
plt.annotate(('Minimum = ' + str(min(y_direction))), (distance_number[5], min(y_direction) + .01))

plt.axhline(y=max(y_direction), color = 'red', linestyle = '--', linewidth = 1.0)
plt.annotate(('Maximum = ' + str(max(y_direction))), (distance_number[5], max(y_direction) + .01))

for i, txt in enumerate(x_direction):
    plt.annotate('{0:.2f}'.format(txt), (distance_number[i] + .25, x_direction[i]))

for i, txt in enumerate(y_direction):
    plt.annotate('{0:.2f}'.format(txt), (distance_number[i] + .25, y_direction[i]))
    
plt.xticks(np.arange(min(distance_number), max(distance_number) + 1, 1.0))

plt.title(label = 'M e a s u r e m e n t      C o n t r o l      C h a r t', fontsize = 40, color = 'green', fontstyle = 'italic', weight = 'bold')

output_dir = input('Measurement control chart destination (Example: C:\chart.jpg): ')
pyplot.savefig(str(output_dir))
plt.show()

# Importing the measurement control chart image into the report file
control_chart_img = openpyxl.drawing.image.Image(str(output_dir))
control_chart_img.anchor = 'O2'
control_chart_img.width = 2400
control_chart_img.height = 800
ws.add_image(control_chart_img)
wb.save(str(xlsx_file))

# Giving summary for the report
print('===========================================================')
print('This is the end of the report. Go to the current folder of \nthis code, and open the report file.')
print('===========================================================')
print('Total number of measurements per each direction:         ' + str(section_distances))
print('Minimum value of the measurements in X-direction:  ' + str("{:.3f}".format(min(x_direction))) + ' mm')
print('Maximum value of the measurements in X-direction:  ' + str("{:.3f}".format(max(x_direction))) + ' mm')
print('Minimum value of the measurements in Y-direction:  ' + str("{:.3f}".format(min(y_direction))) + ' mm')
print('Maximum value of the measurements in Y-direction:  ' + str("{:.3f}".format(max(y_direction))) + ' mm')

